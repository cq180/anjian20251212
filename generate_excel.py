import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

header_font = Font(bold=True, color="FFFFFF", size=11)
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
in_use_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
unused_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
hotkey_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")

def set_header(ws, row, col, value):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = thin_border
    return cell

def set_cell(ws, row, col, value, fill=None, bold=False):
    cell = ws.cell(row=row, column=col, value=value)
    if fill:
        cell.fill = fill
    if bold:
        cell.font = Font(bold=True)
    cell.border = thin_border
    cell.alignment = Alignment(wrap_text=True)
    return cell

# ===== 热键绑定表 =====
ws1 = wb.active
ws1.title = "热键绑定"

ws1.merge_cells('A1:D1')
ws1.cell(row=1, column=1, value="热键绑定 - Hotkey Bindings").font = Font(bold=True, size=14)
ws1.cell(row=1, column=1).alignment = Alignment(horizontal='center')

headers1 = ["快捷键", "脚本路径", "所在目录", "功能说明"]
for col, h in enumerate(headers1, 1):
    set_header(ws1, 2, col, h)

hotkey_data = [
    ("Ctrl+Q", "D:\\game\\按键精灵2014\\screen\\cmd\\tuiyiceng", "根目录", "退1层"),
    ("Alt+Shift+F", "D:\\game\\按键精灵2014\\screen\\cmd\\fengyin", "根目录", "封印"),
    ("Alt+Q", "D:\\game\\按键精灵2014\\screen\\cmd\\tuichuesc", "根目录", "退ESC"),
    ("Ctrl+Alt+J", "D:\\game\\按键精灵2014\\screen\\cmd\\jiejie", "根目录", "结界"),
    ("Ctrl+Alt+H", "D:\\game\\按键精灵2014\\screen\\cmd\\huahezhan", "根目录", "画合战"),
    ("Ctrl+4", "D:\\game\\按键精灵2014\\screen\\cmd\\tui4", "根目录", "退4"),
    ("F11", "[内置] 启动/继续脚本", "-", "系统内置"),
    ("Ctrl+F12", "[内置] 停止脚本", "-", "系统内置"),
]

for i, (key, path, dir, desc) in enumerate(hotkey_data, 3):
    set_cell(ws1, i, 1, key, hotkey_fill, True)
    set_cell(ws1, i, 2, path)
    set_cell(ws1, i, 3, dir)
    set_cell(ws1, i, 4, desc)

ws1.column_dimensions['A'].width = 15
ws1.column_dimensions['B'].width = 55
ws1.column_dimensions['C'].width = 15
ws1.column_dimensions['D'].width = 15

# ===== 按钮面板绑定表 =====
ws2 = wb.create_sheet("按钮绑定")

ws2.merge_cells('A1:F1')
ws2.cell(row=1, column=1, value="按钮面板绑定 - Button Panel Bindings").font = Font(bold=True, size=14)
ws2.cell(row=1, column=1).alignment = Alignment(horizontal='center')

headers2 = ["面板", "按钮位置", "脚本路径", "显示名称", "图片", "状态"]
for col, h in enumerate(headers2, 1):
    set_header(ws2, 2, col, h)

button_data = [
    ("主面板", "0_0", "D:\\game\\按键精灵2014\\screen\\cmd\\qidong", "启动", "有", "使用中"),
    ("主面板", "0_14", "D:\\game\\按键精灵2014\\screen\\cmd\\tuiyiceng", "退1", "", "使用中"),
    ("主面板", "1_14", "D:\\game\\按键精灵2014\\screen\\cmd\\tuichuesc", "退ESC", "", "使用中"),
    ("主面板", "0_13", "D:\\game\\按键精灵2014\\screen\\cmd\\guanbi", "关", "", "使用中"),
    ("主面板", "0_11", "D:\\game\\按键精灵2014\\screen\\cmd\\guanbi", "关启", "", "使用中"),
    ("主面板", "1_11", "D:\\game\\按键精灵2014\\screen\\cmd\\guanbi", "关启", "", "使用中"),
    ("主面板", "2_11", "D:\\game\\按键精灵2014\\screen\\cmd\\guanbi", "关启", "", "使用中"),
    ("主面板", "2_0", "D:\\game\\按键精灵2014\\screen\\cmd\\Record_1_5_1772934168.txt", "放", "有", "使用中"),
    ("主面板", "1_2", "D:\\game\\按键精灵2014\\screen\\cmd\\youjian", "邮件", "有", "使用中"),
    ("主面板", "1_3", "D:\\game\\按键精灵2014\\screen\\cmd\\liao30", "30", "有", "使用中"),
    ("主面板", "1_5", "D:\\game\\按键精灵2014\\screen\\cmd\\每日\\Record_4_2_1771550656.txt", "日2", "", "使用中"),
    ("主面板", "1_6", "D:\\game\\按键精灵2014\\screen\\cmd\\qilin", "麒麟", "有", "使用中"),
    ("主面板", "0_7", "D:\\game\\按键精灵2014\\screen\\cmd\\tuizhi", "周六退治", "有", "使用中"),
    ("主面板", "0_10", "D:\\game\\按键精灵2014\\screen\\cmd\\qidong", "", "有", "使用中"),
    ("主面板", "1_7", "D:\\game\\按键精灵2014\\screen\\cmd\\xiajian1", "1", "有", "使用中"),
    ("主面板", "4_2", "D:\\game\\按键精灵2014\\screen\\cmd\\qiecuoshang", "上", "", "使用中"),
    ("主面板", "4_3", "D:\\game\\按键精灵2014\\screen\\cmd\\qiecuoxia", "下", "", "使用中"),
    ("主面板", "3_0", "D:\\game\\按键精灵2014\\screen\\cmd\\qiandao", "签到", "有", "使用中"),
    ("主面板", "2_1", "D:\\game\\按键精灵2014\\screen\\cmd\\Record_1_6_1772934614.txt", "", "有", "使用中"),
    ("主面板", "1_1", "D:\\game\\按键精灵2014\\screen\\cmd\\地鬼分享", "", "", "使用中"),
    ("主面板", "2_2", "D:\\game\\按键精灵2014\\screen\\cmd\\每日\\白蛋.ks", "白蛋", "有", "使用中"),
    ("主面板", "2_3", "D:\\game\\按键精灵2014\\screen\\cmd\\Record_1_2_1772501452.txt", "", "有", "使用中"),
    ("主面板", "4_1", "D:\\game\\按键精灵2014\\screen\\cmd\\1218\\封印OCR", "OCR", "", "使用中"),
    ("主面板", "4_6", "D:\\game\\按键精灵2014\\screen\\cmd\\每日\\Record_4_3_1771552555.txt", "封任", "", "使用中"),
    ("主面板", "1_0", "D:\\game\\按键精灵2014\\screen\\cmd\\digui", "地鬼", "有", "使用中"),
    ("主面板", "4_0", "D:\\game\\按键精灵2014\\screen\\cmd\\fengyin", "封印", "有", "使用中"),
    ("主面板", "0_1", "D:\\game\\按键精灵2014\\screen\\cmd\\guanbi", "关启", "", "使用中"),
    ("主面板", "0_2", "D:\\game\\按键精灵2014\\screen\\cmd\\庭院自动每日", "", "有", "使用中"),
    ("主面板", "0_4", "D:\\game\\按键精灵2014\\screen\\cmd\\maohouyuan", "猫后院", "有", "使用中"),
    ("主面板", "0_5", "D:\\game\\按键精灵2014\\screen\\cmd\\每日\\Record_4_1_1771465477.txt", "日1", "", "使用中"),
    ("主面板", "0_3", "D:\\game\\按键精灵2014\\screen\\cmd\\每日\\同心兰", "同心兰", "有", "使用中"),
    ("主面板", "3_3", "D:\\game\\按键精灵2014\\screen\\cmd\\huahezhan", "画合战", "有", "使用中"),
    ("主面板", "3_1", "D:\\game\\按键精灵2014\\screen\\cmd\\gouyu", "勾玉", "有", "使用中"),
    ("主面板", "1_10", "D:\\game\\按键精灵2014\\screen\\cmd\\每日\\Record_4_1_1771465477.txt", "日1", "", "使用中"),
    ("备用页P1", "P1_Btn_0_2", "D:\\game\\按键精灵2014\\screen\\cmd\\秘闻\\0每周秘闻-金币6关", "1-6", "有", "使用中"),
    ("备用页P2", "P2_Btn_0_0", "D:\\game\\按键精灵2014\\screen\\cmd\\fujinti", "", "", "使用中"),
    ("备用页P2", "P2_Btn_0_1", "D:\\game\\按键精灵2014\\screen\\cmd\\业原火\\0业原火进入式神录", "", "", "使用中"),
    ("备用页P2", "P2_Btn_0_2", "D:\\game\\按键精灵2014\\screen\\cmd\\dihuazhi", "", "", "使用中"),
    ("备用页P2", "P2_Btn_0_3", "D:\\game\\按键精灵2014\\screen\\cmd\\shangdian", "", "", "使用中"),
    ("备用页P2", "P2_Btn_0_4", "D:\\game\\按键精灵2014\\screen\\cmd\\zhonghuazhi", "", "", "使用中"),
    ("备用页P2", "P2_Btn_0_5", "D:\\game\\按键精灵2014\\screen\\cmd\\zhibo", "", "", "使用中"),
    ("备用页P2", "P2_Btn_0_6", "D:\\game\\按键精灵2014\\screen\\cmd\\战斗\\Record_5_3_1774755713.txt", "", "", "使用中"),
    ("备用页P2", "P2_Btn_0_7", "D:\\game\\按键精灵2014\\screen\\cmd\\1218\\进爬塔", "", "", "使用中"),
    ("备用页P2", "P2_Btn_0_10", "D:\\game\\按键精灵2014\\screen\\cmd\\每日\\Record_4_2_1772383248.txt", "", "", "使用中"),
    ("备用页P2", "P2_Btn_0_11", "D:\\game\\按键精灵2014\\screen\\cmd\\1\\真蛇", "", "", "使用中"),
    ("备用页P2", "P2_Btn_1_0", "D:\\game\\按键精灵2014\\screen\\cmd\\yinjie", "", "", "使用中"),
    ("备用页P2", "P2_Btn_1_1", "D:\\game\\按键精灵2014\\screen\\cmd\\heidan", "", "", "使用中"),
    ("备用页P2", "P2_Btn_1_4", "D:\\game\\按键精灵2014\\screen\\cmd\\weipai", "", "", "使用中"),
    ("备用页P2", "P2_Btn_1_5", "D:\\game\\按键精灵2014\\screen\\cmd\\fengmo", "", "", "使用中"),
    ("备用页P2", "P2_Btn_1_6", "D:\\game\\按键精灵2014\\screen\\cmd\\1218\\合成结界卡", "", "", "使用中"),
    ("备用页P2", "P2_Btn_1_7", "D:\\game\\按键精灵2014\\screen\\cmd\\战斗\\魂30", "", "", "使用中"),
    ("备用页P2", "P2_Btn_2_0", "D:\\game\\按键精灵2014\\screen\\cmd\\jingyanzuduitansuo", "", "", "使用中"),
    ("备用页P2", "P2_Btn_2_1", "D:\\game\\按键精灵2014\\screen\\cmd\\每日\\Record_4_6_1771584544.txt", "", "", "使用中"),
    ("备用页P2", "P2_Btn_2_2", "D:\\game\\按键精灵2014\\screen\\cmd\\战斗\\魂土30领奖", "", "", "使用中"),
    ("备用页P2", "P2_Btn_2_3", "D:\\game\\按键精灵2014\\screen\\cmd\\Record_2_3_1773022117.txt", "", "", "使用中"),
    ("备用页P2", "P2_Btn_2_4", "D:\\game\\按键精灵2014\\screen\\cmd\\1218\\奉纳", "", "", "使用中"),
    ("备用页P2", "P2_Btn_3_0", "D:\\game\\按键精灵2014\\screen\\cmd\\音乐音效", "", "", "使用中"),
    ("备用页P2", "P2_Btn_3_1", "D:\\game\\按键精灵2014\\screen\\cmd\\gongxun", "", "", "使用中"),
    ("备用页P2", "P2_Btn_3_2", "D:\\game\\按键精灵2014\\screen\\cmd\\秘闻\\1每周秘闻-换阵容开战", "", "", "使用中"),
    ("备用页P2", "P2_Btn_3_5", "D:\\game\\按键精灵2014\\screen\\cmd\\1218\\每日全领2602028", "", "", "使用中"),
    ("备用页P2", "P2_Btn_3_6", "D:\\game\\按键精灵2014\\screen\\cmd\\1218\\切磋上下6", "", "", "使用中"),
    ("备用页P2", "P2_Btn_4_0", "D:\\game\\按键精灵2014\\screen\\cmd\\战斗\\Record_4_2_1774438584.txt", "", "", "使用中"),
    ("备用页P2", "P2_Btn_4_1", "D:\\game\\按键精灵2014\\screen\\cmd\\战斗\\进道馆", "", "", "使用中"),
    ("备用页P2", "P2_Btn_4_2", "D:\\game\\按键精灵2014\\screen\\cmd\\1\\真蛇2", "", "", "使用中"),
    ("备用页P2", "P2_Btn_4_4", "D:\\game\\按键精灵2014\\screen\\cmd\\刷30领奖励", "", "", "使用中"),
    ("备用页P2", "P2_Btn_4_5", "D:\\game\\按键精灵2014\\screen\\yuhun30.exe", "", "", "使用中"),
    ("备用页P2", "P2_Btn_4_6", "D:\\game\\按键精灵2014\\screen\\cmd\\寮突破", "", "", "使用中"),
    ("备用页P2", "P2_Btn_4_7", "D:\\game\\按键精灵2014\\screen\\cmd\\每日\\Record_4_3_1772374001.txt", "", "", "使用中"),
]

for i, row_data in enumerate(button_data, 3):
    for col, val in enumerate(row_data, 1):
        fill = in_use_fill if row_data[5] == "使用中" else unused_fill
        set_cell(ws2, i, col, val, fill)

ws2.column_dimensions['A'].width = 12
ws2.column_dimensions['B'].width = 15
ws2.column_dimensions['C'].width = 60
ws2.column_dimensions['D'].width = 12
ws2.column_dimensions['E'].width = 8
ws2.column_dimensions['F'].width = 10

# ===== 未使用脚本表 =====
ws3 = wb.create_sheet("未使用脚本")

ws3.merge_cells('A1:E1')
ws3.cell(row=1, column=1, value="未使用脚本 - Unused Scripts").font = Font(bold=True, size=14)
ws3.cell(row=1, column=1).alignment = Alignment(horizontal='center')

headers3 = ["脚本名称", "完整路径", "类型", "所在目录", "备注"]
for col, h in enumerate(headers3, 1):
    set_header(ws3, 2, col, h)

unused_data = [
    # 根目录
    ("2baidansheng3", "D:\\game\\按键精灵2014\\screen\\cmd\\2baidansheng3", "txt", "根目录", ""),
    ("2baidansheng3man", "D:\\game\\按键精灵2014\\screen\\cmd\\2baidansheng3man", "txt", "根目录", ""),
    ("KYYS", "D:\\game\\按键精灵2014\\screen\\cmd\\KYYS", "文件夹", "根目录", ""),
    ("Messagetofeng", "D:\\game\\按键精灵2014\\screen\\cmd\\Messagetofeng", "未知", "根目录", ""),
    ("Record_1_11_1774433771", "D:\\game\\按键精灵2014\\screen\\cmd\\Record_1_11_1774433771.txt", "txt", "根目录", ""),
    ("Record_4_4_1772501045", "D:\\game\\按键精灵2014\\screen\\cmd\\Record_4_4_1772501045.txt", "txt", "根目录", ""),
    ("Record_4_9_1773146980", "D:\\game\\按键精灵2014\\screen\\cmd\\Record_4_9_1773146980.txt", "txt", "根目录", ""),
    ("Record_5_2_1773821460", "D:\\game\\按键精灵2014\\screen\\cmd\\Record_5_2_1773821460.txt", "txt", "根目录", ""),
    ("baidan", "D:\\game\\按键精灵2014\\screen\\cmd\\baidan", "未知", "根目录", ""),
    ("dihua", "D:\\game\\按键精灵2014\\screen\\cmd\\dihua", "未知", "根目录", ""),
    ("doujiyushe", "D:\\game\\按键精灵2014\\screen\\cmd\\doujiyushe", "未知", "根目录", ""),
    ("fanhe", "D:\\game\\按键精灵2014\\screen\\cmd\\fanhe", "未知", "根目录", ""),
    ("fengmoyushe", "D:\\game\\按键精灵2014\\screen\\cmd\\fengmoyushe", "未知", "根目录", ""),
    ("fengmoyushe2", "D:\\game\\按键精灵2014\\screen\\cmd\\fengmoyushe2", "未知", "根目录", ""),
    ("gongzi", "D:\\game\\按键精灵2014\\screen\\cmd\\gongzi", "未知", "根目录", ""),
    ("huntuqidong", "D:\\game\\按键精灵2014\\screen\\cmd\\huntuqidong", "未知", "根目录", ""),
    ("huntuyushe", "D:\\game\\按键精灵2014\\screen\\cmd\\huntuyushe", "未知", "根目录", ""),
    ("huntugouxuan", "D:\\game\\按键精灵2014\\screen\\cmd\\huntugouxuan", "未知", "根目录", ""),
    ("jiwen", "D:\\game\\按键精灵2014\\screen\\cmd\\jiwen", "未知", "根目录", ""),
    ("jiyang", "D:\\game\\按键精灵2014\\screen\\cmd\\jiyang", "未知", "根目录", ""),
    ("jinbijiacheng", "D:\\game\\按键精灵2014\\screen\\cmd\\jinbijiacheng", "未知", "根目录", ""),
    ("jingyanjinbi", "D:\\game\\按键精灵2014\\screen\\cmd\\jingyanjinbi", "未知", "根目录", ""),
    ("killtupo", "D:\\game\\按键精灵2014\\screen\\cmd\\killtupo", "未知", "根目录", ""),
    ("liaotu", "D:\\game\\按键精灵2014\\screen\\cmd\\liaotu", "未知", "根目录", ""),
    ("qidong", "D:\\game\\按键精灵2014\\screen\\cmd\\qidong", "未知", "根目录", "可能已迁移到1218目录"),
    ("shangdian", "D:\\game\\按键精灵2014\\screen\\cmd\\shangdian", "未知", "根目录", ""),
    ("tansuo28", "D:\\game\\按键精灵2014\\screen\\cmd\\tansuo28", "未知", "根目录", ""),
    ("tansuotui", "D:\\game\\按键精灵2014\\screen\\cmd\\tansuotui", "未知", "根目录", ""),
    ("tongxin", "D:\\game\\按键精灵2014\\screen\\cmd\\tongxin", "未知", "根目录", ""),
    ("youqing", "D:\\game\\按键精灵2014\\screen\\cmd\\youqing", "未知", "根目录", ""),
    ("yuhunjiacheng", "D:\\game\\按键精灵2014\\screen\\cmd\\yuhunjiacheng", "未知", "根目录", ""),
    ("zhibo", "D:\\game\\按键精灵2014\\screen\\cmd\\zhibo", "未知", "根目录", ""),
    # 1218目录
    ("0109活动", "D:\\game\\按键精灵2014\\screen\\cmd\\1218\\0109活动", "未知", "1218", ""),
    ("123", "D:\\game\\按键精灵2014\\screen\\cmd\\1218\\123", "未知", "1218", ""),
    ("qidong", "D:\\game\\按键精灵2014\\screen\\cmd\\1218\\qidong", "未知", "1218", ""),
    ("切磋上下", "D:\\game\\按键精灵2014\\screen\\cmd\\1218\\切磋上下", "未知", "1218", ""),
    ("切磋上下12", "D:\\game\\按键精灵2014\\screen\\cmd\\1218\\切磋上下12", "未知", "1218", ""),
    ("刷镰鼬", "D:\\game\\按键精灵2014\\screen\\cmd\\1218\\刷镰鼬", "未知", "1218", ""),
    ("封印OCR1", "D:\\game\\按键精灵2014\\screen\\cmd\\1218\\封印OCR1", "未知", "1218", ""),
    ("封印完成", "D:\\game\\按键精灵2014\\screen\\cmd\\1218\\封印完成", "未知", "1218", ""),
    ("庭院自动", "D:\\game\\按键精灵2014\\screen\\cmd\\1218\\庭院自动", "未知", "1218", ""),
    ("日常", "D:\\game\\按键精灵2014\\screen\\cmd\\1218\\日常", "未知", "1218", ""),
    ("每日全领", "D:\\game\\按键精灵2014\\screen\\cmd\\1218\\每日全领", "未知", "1218", ""),
    ("逢魔阴界周五", "D:\\game\\按键精灵2014\\screen\\cmd\\1218\\逢魔阴界周五", "未知", "1218", ""),
    # 活动目录
    ("Record_1_1_1771332588", "D:\\game\\按键精灵2014\\screen\\cmd\\活动\\Record_1_1_1771332588.txt", "txt", "活动", ""),
    ("小白鬼王活动", "D:\\game\\按键精灵2014\\screen\\cmd\\活动\\小白鬼王活动", "未知", "活动", ""),
    ("爬塔预设", "D:\\game\\按键精灵2014\\screen\\cmd\\活动\\爬塔预设", "未知", "活动", ""),
    ("葛叶", "D:\\game\\按键精灵2014\\screen\\cmd\\活动\\葛叶", "未知", "活动", ""),
    # 秘闻目录
    ("12312", "D:\\game\\按键精灵2014\\screen\\cmd\\秘闻\\12312", "未知", "秘闻", ""),
    ("1每周秘闻-分享", "D:\\game\\按键精灵2014\\screen\\cmd\\秘闻\\1每周秘闻-分享", "未知", "秘闻", ""),
    ("1每周秘闻-继续打", "D:\\game\\按键精灵2014\\screen\\cmd\\秘闻\\1每周秘闻-继续打", "未知", "秘闻", ""),
    ("2金币关", "D:\\game\\按键精灵2014\\screen\\cmd\\秘闻\\2金币关", "未知", "秘闻", ""),
    # 每日目录
    ("Record_1_4_1771424496", "D:\\game\\按键精灵2014\\screen\\cmd\\每日\\Record_1_4_1771424496.txt", "txt", "每日", ""),
    ("Record_2_6_1771584797", "D:\\game\\按键精灵2014\\screen\\cmd\\每日\\Record_2_6_1771584797.txt", "txt", "每日", ""),
    ("Record_2_7_1771562142", "D:\\game\\按键精灵2014\\screen\\cmd\\每日\\Record_2_7_1771562142.txt", "txt", "每日", ""),
    ("Record_3_2_1772368884", "D:\\game\\按键精灵2014\\screen\\cmd\\每日\\Record_3_2_1772368884.txt", "txt", "每日", ""),
    ("Record_4_4_1771557646", "D:\\game\\按键精灵2014\\screen\\cmd\\每日\\Record_4_4_1771557646.txt", "txt", "每日", ""),
    # 战斗目录
    ("huntu0", "D:\\game\\按键精灵2014\\screen\\cmd\\战斗\\huntu0", "未知", "战斗", ""),
    # 业原火目录
    ("1继续", "D:\\game\\按键精灵2014\\screen\\cmd\\业原火\\1继续", "未知", "业原火", ""),
    ("停止换蛋", "D:\\game\\按键精灵2014\\screen\\cmd\\业原火\\停止换蛋", "未知", "业原火", ""),
    ("金币经验加成-主", "D:\\game\\按键精灵2014\\screen\\cmd\\业原火\\金币经验加成-主", "未知", "业原火", ""),
    # 1目录
    ("合碎片", "D:\\game\\按键精灵2014\\screen\\cmd\\1\\合碎片", "未知", "1", ""),
    ("放红蛋", "D:\\game\\按键精灵2014\\screen\\cmd\\1\\放红蛋", "未知", "1", ""),
    ("画质-中", "D:\\game\\按键精灵2014\\screen\\cmd\\1\\画质-中", "未知", "1", ""),
    ("画质-低", "D:\\game\\按键精灵2014\\screen\\cmd\\1\\画质-低", "未知", "1", ""),
    ("道馆预设", "D:\\game\\按键精灵2014\\screen\\cmd\\1\\道馆预设", "未知", "1", ""),
    ("音乐音效", "D:\\game\\按键精灵2014\\screen\\cmd\\1\\音乐音效", "未知", "1", ""),
]

for i, row_data in enumerate(unused_data, 3):
    for col, val in enumerate(row_data, 1):
        set_cell(ws3, i, col, val, unused_fill)

ws3.column_dimensions['A'].width = 25
ws3.column_dimensions['B'].width = 60
ws3.column_dimensions['C'].width = 10
ws3.column_dimensions['D'].width = 12
ws3.column_dimensions['E'].width = 25

# ===== 统计表 =====
ws4 = wb.create_sheet("统计汇总")

ws4.merge_cells('A1:B1')
ws4.cell(row=1, column=1, value="脚本使用统计 - Summary").font = Font(bold=True, size=14)
ws4.cell(row=1, column=1).alignment = Alignment(horizontal='center')

summary_data = [
    ("项目", "数量"),
    ("热键绑定脚本数", 6),
    ("按钮绑定脚本数", 67),
    ("未使用脚本数", len(unused_data)),
    ("脚本总数", 6 + 67 + len(unused_data)),
    ("", ""),
    ("目录统计", ""),
    ("根目录脚本数", 32),
    ("1218目录脚本数", 12),
    ("活动目录脚本数", 4),
    ("秘闻目录脚本数", 4),
    ("每日目录脚本数", 5),
    ("战斗目录脚本数", 1),
    ("业原火目录脚本数", 3),
    ("1目录脚本数", 6),
]

for i, (col1, col2) in enumerate(summary_data, 3):
    set_cell(ws4, i, 1, col1, bold=True)
    set_cell(ws4, i, 2, col2)

ws4.column_dimensions['A'].width = 25
ws4.column_dimensions['B'].width = 15

wb.save("D:\\game\\按键精灵2014\\screen\\cmd\\脚本使用情况分析.xlsx")
print(f"Excel文件已生成: D:\\game\\按键精灵2014\\screen\\cmd\\脚本使用情况分析.xlsx")
